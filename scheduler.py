import pandas as pd
from datetime import datetime, timedelta
import pulp
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
import calendar
import random


class InternScheduler:
    def __init__(self, start_date, end_date, interns, units_per_intern, min_interns_per_duty, minimum_spacing=1):
        self.start_date = start_date
        self.end_date = end_date
        self.interns = interns
        self.units_per_intern = units_per_intern
        self.min_interns_per_duty = min_interns_per_duty
        self.minimum_spacing = minimum_spacing
        self.dates = self.generate_dates(start_date, end_date)
        self.schedule = None
        self.intern_metrics = None

    def generate_dates(self, start_date, end_date):
        return pd.date_range(start=start_date, end=end_date).to_pydatetime().tolist()

    def calculate_total_required_units(self):
        total_units = 0
        for d in self.dates:
            if d.weekday() == 5:  # Saturday
                total_units += 2
            elif d.weekday() == 6:  # Sunday
                total_units += 3
            else:  # Weekdays
                total_units += 1
        return total_units

    def verify_units(self, total_required_units):
        total_required_units_interns = total_required_units * self.min_interns_per_duty
        total_intern_units = sum(self.units_per_intern.values())

        if total_intern_units < total_required_units:
            raise ValueError(f"Total number of intern units ({total_intern_units}) is insufficient to cover the required {total_required_units} units.")
        elif total_intern_units < total_required_units_interns:
            raise ValueError(f"Total number of intern units ({total_intern_units}) is insufficient to cover the minimum {self.min_interns_per_duty} interns per duty. The number of total units needed is ({total_required_units_interns}).")
        else:
            print(f"Total number of intern units ({total_intern_units}) is sufficient to cover the required {total_required_units} units and the minimum {self.min_interns_per_duty} interns per duty.")

    def solve(self, randomize = True):
        total_required_units = self.calculate_total_required_units()
        self.verify_units(total_required_units)

        # Create shuffled list for randomness
        shuffled_interns = self.interns[:]  # Create a copy of the original list
        random.shuffle(shuffled_interns)

        if randomize:
            random.shuffle(self.interns)

        # Initialize the problem
        prob = pulp.LpProblem("Intern_Schedule_Optimization", pulp.LpMaximize)

        # Define decision variables (1 if an intern is assigned to a day, 0 otherwise)
        intern_vars = pulp.LpVariable.dicts("Duty", ((i, d) for i in self.interns for d in self.dates), cat='Binary')

        # Add constraints to ensure the minimum number of interns per duty
        for d in self.dates:
            prob += pulp.lpSum([intern_vars[(i, d)] for i in self.interns]) >= self.min_interns_per_duty, f"Min_Interns_on_{d}"

        # Add constraints to ensure that each intern fills the required number of units
        for i in self.interns:
            total_units = pulp.lpSum([intern_vars[(i, d)] * self.get_units_for_day(d) for d in self.dates])
            prob += total_units == self.units_per_intern[i], f"Units_for_{i}"
        
        # Add constraints to ensure that each intern is given the same amount of weekends
        # Calculate the total number of weekdays, Saturdays, and Sundays
        num_weekdays = sum(1 for d in self.dates if d.weekday() < 5)
        num_saturdays = sum(1 for d in self.dates if d.weekday() == 5)
        num_sundays = sum(1 for d in self.dates if d.weekday() == 6)

        # Calculate the expected number of duties per intern
        expected_weekdays = (num_weekdays * self.min_interns_per_duty) // len(self.interns)
        expected_saturdays = (num_saturdays * self.min_interns_per_duty) // len(self.interns)
        expected_sundays = (num_sundays * self.min_interns_per_duty) // len(self.interns)

        # Calculate the remainder of duties that need to be distributed
        remainder_weekdays = (num_weekdays * self.min_interns_per_duty) % len(self.interns)
        remainder_saturdays = (num_saturdays * self.min_interns_per_duty) % len(self.interns)
        remainder_sundays = (num_sundays * self.min_interns_per_duty) % len(self.interns)

        weekend_distribution = {intern : [expected_saturdays, expected_sundays] for intern in self.interns}

        while remainder_saturdays > 0 and remainder_sundays > 0:
          for intern in shuffled_interns:
            if remainder_saturdays == 0 and remainder_sundays == 0:
              break
            if remainder_saturdays > 0:
              weekend_distribution[intern][0] += 1
              remainder_saturdays -= 1
              continue
            if remainder_sundays > 0:
              weekend_distribution[intern][1] += 1
              remainder_sundays -= 1
              continue

        for i in self.interns:
          # Constraint for Saturdays
          prob += pulp.lpSum([intern_vars[(i, d)] for d in self.dates if d.weekday() == 5]) == weekend_distribution[i][0], f"Saturday_Duties_for_{i}"

          # Constraint for Sundays
          prob += pulp.lpSum([intern_vars[(i, d)] for d in self.dates if d.weekday() == 6]) == weekend_distribution[i][1], f"Sunday_Duties_for_{i}"

        # Enforce minimum spacing between shifts of the same intern
        for i in self.interns:
            for j in range(len(self.dates)):  # Loop through all dates
                d1 = self.dates[j]
                for k in range(1, self.minimum_spacing + 1):  # Ensure spacing up to the minimum_spacing days
                    if j + k < len(self.dates):  # Ensure we are within the valid date range
                        d2 = self.dates[j + k]
                        prob += intern_vars[(i, d1)] + intern_vars[(i, d2)] <= 1, f"Min_Spacing_Shifts_{i}_{d1}_{d2}"

        # Solve the problem
        prob.solve()

        # Extract the solution and calculate metrics for each intern
        self.schedule = {d: [] for d in self.dates}
        self.intern_metrics = {}

        for i in self.interns:
            shifts = []
            total_units_used = 0
            weekdays = saturdays = sundays = 0

            for d in self.dates:
                if pulp.value(intern_vars[(i, d)]) == 1:
                    shifts.append(d)
                    total_units_used += self.get_units_for_day(d)
                    self.schedule[d].append(i)

                    # Count the number of weekdays, saturdays, and sundays
                    if d.weekday() < 5:
                        weekdays += 1
                    elif d.weekday() == 5:
                        saturdays += 1
                    elif d.weekday() == 6:
                        sundays += 1

            if shifts:
                # Calculate average spacing between shifts
                total_spacing = (shifts[0] - self.dates[0]).days + (self.dates[-1] - shifts[-1]).days
                for j in range(1, len(shifts)):
                    total_spacing += (shifts[j] - shifts[j-1]).days

                average_spacing = total_spacing / (len(shifts) + 1)
            else:
                average_spacing = None  # No shifts assigned

            # Store metrics
            self.intern_metrics[i] = {
                "total_units_used": total_units_used,
                "average_spacing": average_spacing,
                "num_shifts": len(shifts),
                "weekdays": weekdays,
                "saturdays": saturdays,
                "sundays": sundays
            }

    def get_units_for_day(self, date):
        if date.weekday() == 5:  # Saturday
            return 2
        elif date.weekday() == 6:  # Sunday
            return 3
        else:
            return 1

    def print_schedule(self):
        if self.schedule:
            print("Schedule:")
            for day, interns_on_duty in self.schedule.items():
                print(f"{day.date()} : {interns_on_duty}")
        else:
            print("No schedule generated. Please run the `.solve()` method.")

    def print_intern_metrics(self):
        if self.intern_metrics:
            print("\nIntern Metrics:")
            for intern, metrics in self.intern_metrics.items():
                print(f"{intern}: Total Units Used: {metrics['total_units_used']}, Average Spacing: {metrics['average_spacing']}, "
                      f"Number of Shifts: {metrics['num_shifts']}, Weekdays: {metrics['weekdays']}, "
                      f"Saturdays: {metrics['saturdays']}, Sundays: {metrics['sundays']}")
        else:
            print("No metrics available. Please run the `.solve()` method.")

    def create_calendar(self, output_file, row_format=True):
        if self.schedule is None:
            raise ValueError("No schedule found. Please run the `.solve()` method first.")

        if not row_format:
            # Create a list of all dates between start_date and end_date
            dates = pd.date_range(start=self.start_date, end=self.end_date)

            # Initialize the calendar DataFrame
            calendar_data = []

            # Track the current week
            current_week = []

            # Iterate over each date in the range
            for date in dates:
                # Get the interns scheduled for the current date
                interns_on_duty = self.schedule.get(date.to_pydatetime(), [])

                # Format the date and interns as a string
                day_entry = f"{date.strftime('%Y-%m-%d')}:\n" + ', '.join(interns_on_duty) if interns_on_duty else "No duty"

                # Add the entry to the current week list
                current_week.append(day_entry)

                # Check if we have filled a week (7 days) or it's the last day
                if len(current_week) == 7 or date == dates[-1]:
                    # Add the week to the calendar data
                    calendar_data.append(current_week)
                    # Reset the current week list
                    current_week = []

            # Convert the calendar data to a DataFrame
            calendar_df = pd.DataFrame(calendar_data, columns=['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday'])

            # Create a workbook and add the DataFrame as a sheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Intern Duty Calendar"

            # Write the DataFrame to the Excel sheet
            for row in dataframe_to_rows(calendar_df, index=False, header=True):
                ws.append(row)

            # Save the workbook to an Excel file
            wb.save(output_file)
            print(f"Calendar saved to {output_file}")

        else:
            # Create a list of all dates between start_date and end_date
            dates = pd.date_range(start=self.start_date, end=self.end_date)

            # Define a function to calculate the number of units for the day based on the day of the week
            def get_units_for_day(date):
                if date.weekday() == 5:  # Saturday
                    return 2
                elif date.weekday() == 6:  # Sunday
                    return 3
                else:
                    return 1

            # Initialize the calendar data for rows
            calendar_data = []

            # Track the maximum number of interns on any day
            max_interns_per_day = 0

            # Iterate over each date in the range
            for date in dates:
                # Get the interns scheduled for the current date
                interns_on_duty = self.schedule.get(date.to_pydatetime(), [])

                # Update the maximum number of interns per day
                max_interns_per_day = max(max_interns_per_day, len(interns_on_duty))

                # Format the full date
                full_date = date.strftime('%A, %B %d, %Y')

                # Get the number of units for the day
                units = get_units_for_day(date)

                # Create a row with the date, units, and intern list
                row = [full_date, units] + interns_on_duty

                # Add the row to the calendar data
                calendar_data.append(row)

            # Prepare column headers: Date, Units, and "Intern 1", "Intern 2", etc.
            columns = ['Date', 'Units'] + [f"Intern {i+1}" for i in range(max_interns_per_day)]

            # Convert the calendar data to a DataFrame
            calendar_df = pd.DataFrame(calendar_data, columns=columns)

            # Fill missing intern slots with "No duty"
            calendar_df.fillna(" ", inplace=True)

            # Create a workbook and add the DataFrame as a sheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Intern Duty Calendar"

            # Hide gridlines in the sheet
            ws.sheet_view.showGridLines = False

            # Write the DataFrame to the Excel sheet
            for row in dataframe_to_rows(calendar_df, index=False, header=True):
                ws.append(row)

            # Autofit columns
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter  # Get the column name
                for cell in col:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

            # Set a minimum column width of 150 pixels (~21.43 Excel units)
            minimum_column_width = 21.43
            for col in ws.columns:
                column = col[0].column_letter  # Get the column letter
                ws.column_dimensions[column].width = max(ws.column_dimensions[column].width, minimum_column_width)

            # Apply formatting to headers: bold text
            header_font = Font(bold=True)
            for cell in ws[1]:
                cell.font = header_font

            # Apply borders to all cells
            thin_border = Border(left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin'))

            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border

            # Center the units column
            for cell in ws['B']:  # Column B is the "Units" column
                cell.alignment = Alignment(horizontal="center")

            # Assign a color to each intern and create fills
            intern_colors = {}

            # Extended color palette with 15 distinct colors
            color_choices = [
                'FF9999', '99FF99', '9999FF', 'FFFF99', 'FF99FF', '99FFFF', 'FFCC99',
                'FF6666', '66FF66', '6666FF', 'FF66FF', '66FFFF', 'CCCC66', 'FF9966', '66CCFF'
            ]

            # Make sure we shuffle colors so that each run can have different colors
            random.shuffle(color_choices)

            # Collect all unique intern names to assign colors
            all_interns = set()
            for row in calendar_data:
                all_interns.update(row[2:])  # Collect all intern names (skipping date and units)

            # Assign unique colors to each intern
            for index, intern in enumerate(all_interns):
                color = color_choices[index % len(color_choices)]
                intern_colors[intern] = PatternFill(start_color=color, end_color=color, fill_type="solid")

            # Apply colors to intern cells in the sheet
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=2 + max_interns_per_day):
                for cell in row:
                    intern_name = cell.value
                    if intern_name in intern_colors:
                        cell.fill = intern_colors[intern_name]

            # Save the workbook to an Excel file
            wb.save(output_file)
            print(f"Calendar saved to {output_file}")